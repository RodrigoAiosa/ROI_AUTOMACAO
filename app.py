import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage


# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ROI de Automação",
    page_icon="🦉",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Syne', sans-serif; }
.stApp { background: #0a0a0f; color: #e8e8f0; }
section[data-testid="stSidebar"] { background: #0f0f1a !important; border-right: 1px solid #1e1e2e; }
h1, h2, h3 { font-family: 'Syne', sans-serif !important; font-weight: 800 !important; }

section[data-testid="stSidebar"] label { color: #9ca3af !important; font-size: 12px !important; }
section[data-testid="stSidebar"] .stSelectbox > div > div { background: #1a1a2e !important; border: 1px solid #2d2d4e !important; color: #e8e8f0 !important; }
section[data-testid="stSidebar"] .stNumberInput input { background: #1a1a2e !important; color: #e8e8f0 !important; border: 1px solid #2d2d4e !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; }

.stNumberInput input { background: #1a1a2e !important; color: #e8e8f0 !important; border: 1px solid #2d2d4e !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; }
.stNumberInput input:focus { border-color: #4ade80 !important; box-shadow: 0 0 0 2px rgba(74,222,128,0.15) !important; }
label { color: #9ca3af !important; font-size: 13px !important; }
.stSlider > div > div > div { background: #1e1e2e !important; }

.metric-card { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); border: 1px solid #2d2d4e; border-radius: 12px; padding: 18px 14px; text-align: center; transition: transform 0.2s, border-color 0.2s; overflow: hidden; }
.metric-card:hover { transform: translateY(-3px); border-color: #4ade80; }
.metric-label { font-family: 'Space Mono', monospace; font-size: clamp(8px, 0.8vw, 11px); text-transform: uppercase; letter-spacing: 1.5px; color: #6b7280; margin-bottom: 10px; white-space: nowrap; }
.metric-value { font-family: 'Syne', sans-serif; font-size: clamp(16px, 2vw, 26px); font-weight: 800; color: #4ade80; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; display: block; }
.metric-value.warning { color: #fb923c; }
.metric-value.info    { color: #60a5fa; }
.metric-value.danger  { color: #f87171; }

.header-tag { font-family: 'Space Mono', monospace; font-size: 11px; color: #4ade80; letter-spacing: 3px; text-transform: uppercase; margin-bottom: 4px; }
.section-title { font-family: 'Space Mono', monospace; font-size: 12px; letter-spacing: 2px; text-transform: uppercase; color: #4ade80; border-left: 3px solid #4ade80; padding-left: 10px; margin: 20px 0 14px 0; }

.summary-box { background: linear-gradient(135deg, #0d2818 0%, #0a1628 100%); border: 1px solid #4ade80; border-radius: 12px; padding: 22px; margin-top: 20px; }
.summary-box p { font-family: 'Space Mono', monospace; font-size: 13px; color: #d1fae5; line-height: 1.8; margin: 0; }

.scenario-badge { display: inline-block; background: #1a1a2e; border: 1px solid #4ade80; border-radius: 20px; padding: 3px 14px; font-family: 'Space Mono', monospace; font-size: 11px; color: #4ade80; letter-spacing: 1px; margin-bottom: 8px; }

div[data-testid="stDownloadButton"] button { background: linear-gradient(135deg, #166534, #14532d) !important; color: #4ade80 !important; border: 1px solid #4ade80 !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; font-size: 12px !important; letter-spacing: 1px !important; padding: 10px 20px !important; width: 100% !important; transition: all 0.2s !important; }
div[data-testid="stDownloadButton"] button:hover { background: #4ade80 !important; color: #0a0a0f !important; }
</style>
""", unsafe_allow_html=True)

# ── Cenários pré-definidos ─────────────────────────────────────────────────────
CENARIOS = {
    "🎯 Personalizado": None,
    "🤖 Automação de Relatórios": {"custo_dev": 3000.0, "custo_manut": 150.0, "horas_mes": 20.0, "valor_hora": 60.0, "anos": 2},
    "📧 Disparo de E-mails":      {"custo_dev": 1500.0, "custo_manut": 50.0,  "horas_mes": 15.0, "valor_hora": 40.0, "anos": 1},
    "🔄 Integração ETL":           {"custo_dev": 8000.0, "custo_manut": 500.0, "horas_mes": 60.0, "valor_hora": 80.0, "anos": 3},
    "📊 Scraping de Dados":        {"custo_dev": 2500.0, "custo_manut": 100.0, "horas_mes": 30.0, "valor_hora": 55.0, "anos": 2},
    "🧾 Emissão de NF-e":          {"custo_dev": 5000.0, "custo_manut": 200.0, "horas_mes": 44.0, "valor_hora": 50.0, "anos": 3},
    "📁 Organização de Arquivos":  {"custo_dev": 800.0,  "custo_manut": 30.0,  "horas_mes": 8.0,  "valor_hora": 35.0, "anos": 1},
}

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="header-tag">// Cenários</div>', unsafe_allow_html=True)
    st.markdown("## Configuração")
    st.markdown("Selecione um cenário ou configure manualmente.")
    st.markdown("---")

    cenario_sel = st.selectbox("📋 Cenário", list(CENARIOS.keys()))
    dados = CENARIOS[cenario_sel]

    st.markdown("---")
    st.markdown('<div class="section-title">Custos</div>', unsafe_allow_html=True)
    custo_dev   = st.number_input("Custo de desenvolvimento (R$)", min_value=0.0, value=float(dados["custo_dev"]) if dados else 2000.0, step=100.0)
    custo_manut = st.number_input("Manutenção mensal (R$)",         min_value=0.0, value=float(dados["custo_manut"]) if dados else 200.0,  step=50.0)

    st.markdown('<div class="section-title">Benefícios</div>', unsafe_allow_html=True)
    horas_mes  = st.number_input("Horas manuais economizadas/mês", min_value=0.0, value=float(dados["horas_mes"]) if dados else 44.0, step=1.0)
    valor_hora = st.number_input("Valor/hora do profissional (R$)", min_value=0.0, value=float(dados["valor_hora"]) if dados else 50.0, step=5.0)
    anos       = st.slider("Período de análise (anos)", 1, 5, value=int(dados["anos"]) if dados else 1)

    st.markdown("---")

# ── Cálculos ───────────────────────────────────────────────────────────────────
meses_total   = anos * 12
benef_mensal  = horas_mes * valor_hora
benef_total   = benef_mensal * meses_total
custo_total   = custo_dev + (custo_manut * meses_total)
lucro_liquido = benef_total - custo_total
roi           = ((benef_total - custo_total) / custo_total * 100) if custo_total > 0 else 0
econ_mensal   = benef_mensal - custo_manut
payback       = (custo_dev / econ_mensal) if econ_mensal > 0 else float('inf')

# Cálculo da representatividade da manutenção
perc_manut = (custo_manut / benef_mensal * 100) if benef_mensal > 0 else 0

# ── Formatação de payback legível ──────────────────────────────────────────────
def fmt_payback(payback_meses, modo="longo"):
    if payback_meses == float('inf'):
        return "indefinido" if modo == "longo" else "∞"
    dias_total = payback_meses * 30
    if dias_total < 1:
        return "menos de 1 dia" if modo == "longo" else "< 1d"
    if dias_total < 30:
        dias = round(dias_total)
        return f"{dias} dia{'s' if dias > 1 else ''}" if modo == "longo" else f"{dias}d"
    meses = int(payback_meses)
    dias  = round((payback_meses - meses) * 30)
    if dias >= 30:
        meses += 1
        dias   = 0
    if modo == "curto":
        return f"{meses}m {dias}d" if dias > 0 else f"{meses}m"
    partes = []
    if meses > 0:
        partes.append(f"{meses} {'mês' if meses == 1 else 'meses'}")
    if dias > 0:
        partes.append(f"{dias} {'dia' if dias == 1 else 'dias'}")
    return " e ".join(partes)

payback_texto = fmt_payback(payback, modo="longo")
payback_curto = fmt_payback(payback, modo="curto")

# ── Textos de resumo para Excel ─────────────────────────────────────────────
veredicto_excel = ("Excelente investimento" if roi > 200 else
                   "Bom investimento"       if roi > 50  else
                   "Marginal"               if roi > 0   else
                   "Prejuizo - revise os parametros.")

veredicto_emoji = ("✅ Excelente investimento." if roi > 200 else
                   "✅ Bom investimento."       if roi > 50  else
                   "⚠️ Investimento marginal."  if roi > 0   else
                   "❌ Prejuízo — revise os parâmetros.")

def fmt_brl(v):
    s, a = ("-" if v < 0 else ""), abs(v)
    if a >= 1_000_000: return f"{s}R$ {a/1_000_000:.1f}M"
    if a >= 1_000:     return f"{s}R$ {a/1_000:.1f}k"
    return f"{s}R$ {a:.0f}"

def fmt_roi(v):
    s, a = ("-" if v < 0 else ""), abs(v)
    if a >= 1_000: return f"{s}{a/1_000:.1f}k%"
    return f"{s}{a:.0f}%"

def cor_card(v, pos="metric-value", neg="metric-value danger"):
    return pos if v >= 0 else neg

# ── Gerar Excel ────────────────────────────────────────────────────────────────
def gerar_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    thin = Side(style="thin", color="2D2D4E")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, txt, bg="FF1A1A2E", fc="FF4ADE80", bold=True, sz=11):
        cell.value = txt
        cell.font = Font(name="Arial", bold=bold, color=fc, size=sz)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda

    def val_cell(cell, v, fmt=None, fc="FFE8E8F0", bold=False, bg="FF0F0F1A"):
        cell.value = v
        cell.font = Font(name="Arial", bold=bold, color=fc, size=10)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
        if fmt: cell.number_format = fmt

    ws.merge_cells("A1:F1")
    ws["A1"].value = "CALCULADORA DE ROI — AUTOMAÇÃO PYTHON"
    ws["A1"].font = Font(name="Arial", bold=True, color="FF4ADE80", size=14)
    ws["A1"].fill = PatternFill("solid", start_color="FF0A0A0F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Cenário: {cenario_sel}"
    ws["A2"].font = Font(name="Arial", italic=True, color="FF9CA3AF", size=10)
    ws["A2"].fill = PatternFill("solid", start_color="FF0A0A0F")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A4:F4")
    hdr(ws["A4"], "PARÂMETROS DE ENTRADA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    for i, h in enumerate(["Parâmetro", "Valor", "Unidade", "", "Parâmetro", "Valor"], 1):
        hdr(ws.cell(5, i), h, sz=9)
    
    params = [
        ("Custo de Desenvolvimento", custo_dev,   "R$",    "Horas economizadas/mês",  horas_mes,  "h/mês"),
        ("Manutenção mensal",        custo_manut, "R$/mês","Valor/hora profissional",  valor_hora, "R$/h"),
        ("Período de análise",       anos,        "anos",  "",                         "",         ""),
    ]
    for i, (l1, v1, u1, l2, v2, u2) in enumerate(params, 6):
        val_cell(ws.cell(i, 1), l1, fc="FF9CA3AF", bg="FF1A1A2E")
        val_cell(ws.cell(i, 2), v1, fmt='"R$"#,##0.00' if isinstance(v1, float) else "0", fc="FFE8E8F0", bold=True, bg="FF16213E")
        val_cell(ws.cell(i, 3), u1, fc="FF6B7280", bg="FF1A1A2E")
        val_cell(ws.cell(i, 4), "", bg="FF0A0A0F")
        val_cell(ws.cell(i, 5), l2, fc="FF9CA3AF", bg="FF1A1A2E")
        val_cell(ws.cell(i, 6), v2 if v2 != "" else None, fmt="#,##0.00" if isinstance(v2, float) else None, fc="FFE8E8F0", bold=bool(v2), bg="FF16213E" if v2 != "" else "FF0A0A0F")

    ws.merge_cells("A10:F10")
    hdr(ws["A10"], "RESULTADOS", bg="FF0F2818", fc="FF4ADE80", sz=10)
    for i, h in enumerate(["Métrica", "Valor", "Observação"], 1):
        hdr(ws.cell(11, i), h, sz=9)

    veredicto = ("Excelente investimento" if roi > 200 else "Bom investimento" if roi > 50 else "Marginal" if roi > 0 else "Prejuizo")
    resultados = [
        ("Benefício Mensal",   benef_mensal,  '"R$"#,##0.00', "Horas × Valor/hora"),
        ("Custo Total",        custo_total,   '"R$"#,##0.00', "Dev + Manutenção acumulada"),
        ("Benefício Total",    benef_total,   '"R$"#,##0.00', f"Acumulado em {anos} ano(s)"),
        ("Lucro Líquido",      lucro_liquido, '"R$"#,##0.00', "Benefício Total − Custo Total"),
        (f"ROI ({anos}a)",      roi / 100,      '0.00%',        veredicto),
        ("Payback",            payback_texto, None,            "Tempo para recuperar o investimento"),
    ]
    for i, (lbl, v, fmt, obs) in enumerate(resultados, 12):
        val_cell(ws.cell(i, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
        cor = "FF4ADE80" if (isinstance(v, (int, float)) and v >= 0) else "FFF87171"
        val_cell(ws.cell(i, 2), v, fmt=fmt, fc=cor, bold=True, bg="FF16213E")
        val_cell(ws.cell(i, 3), obs, fc="FF9CA3AF", bg="FF1A1A2E")

    for col, w in zip("ABCDEF", [28, 16, 30, 2, 28, 14]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Sidebar: exportar ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
    excel_buf = gerar_excel()
    nome_arquivo = cenario_sel.replace(" ", "_").replace("/", "").replace("🎯","").replace("🤖","").replace("📧","").replace("🔄","").replace("📊","").replace("🧾","").replace("📁","").strip()
    st.download_button(
        label="📥 Exportar para Excel",
        data=excel_buf,
        file_name=f"roi_{nome_arquivo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown("<br>", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown('<div class="header-tag">// Calculadora</div>', unsafe_allow_html=True)
st.title("Calculadora ROI de Automação")
if cenario_sel != "🎯 Personalizado":
    st.markdown(f'<span class="scenario-badge">{cenario_sel}</span>', unsafe_allow_html=True)
st.caption("Descubra em quanto tempo sua automação se paga — e o quanto ela rende.")
st.markdown("---")

# ── Cards ──────────────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Resultados</div>', unsafe_allow_html=True)
c1, c2, c3, c4, c5 = st.columns(5)
pay_cor = "metric-value" if payback < 12 else "metric-value warning"
with c1:
    st.markdown(f'<div class="metric-card"><div class="metric-label">Benefício Mensal</div><div class="metric-value info">{fmt_brl(benef_mensal)}</div></div>', unsafe_allow_html=True)
with c2:
    st.markdown(f'<div class="metric-card"><div class="metric-label">Payback</div><div class="{pay_cor}">{payback_curto}</div></div>', unsafe_allow_html=True)
with c3:
    roi_cor = cor_card(roi, "metric-value", "metric-value danger")
    st.markdown(f'<div class="metric-card"><div class="metric-label">ROI ({anos}a)</div><div class="{roi_cor}">{fmt_roi(roi)}</div></div>', unsafe_allow_html=True)
with c4:
    ll_cor = cor_card(lucro_liquido, "metric-value", "metric-value danger")
    st.markdown(f'<div class="metric-card"><div class="metric-label">Lucro Líquido</div><div class="{ll_cor}">{fmt_brl(lucro_liquido)}</div></div>', unsafe_allow_html=True)
with c5:
    st.markdown(f'<div class="metric-card"><div class="metric-label">Custo Total</div><div class="metric-value warning">{fmt_brl(custo_total)}</div></div>', unsafe_allow_html=True)

# ── Gráfico ────────────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Projeção Acumulada</div>', unsafe_allow_html=True)
meses_eixo = list(range(0, meses_total + 1))
custo_acum = [custo_dev + custo_manut * m for m in meses_eixo]
benef_acum = [benef_mensal * m for m in meses_eixo]
saldo_acum = [b - c for b, c in zip(benef_acum, custo_acum)]

fig = go.Figure()
fig.add_trace(go.Scatter(x=meses_eixo, y=benef_acum, name="Benefício acumulado", line=dict(color="#4ade80", width=2.5)))
fig.add_trace(go.Scatter(x=meses_eixo, y=custo_acum, name="Custo acumulado",     line=dict(color="#f87171", width=2.5, dash="dash")))
fig.add_trace(go.Scatter(x=meses_eixo, y=saldo_acum, name="Saldo líquido",       line=dict(color="#60a5fa", width=2), fill="tozeroy", fillcolor="rgba(96,165,250,0.08)"))
fig.add_hline(y=0, line_dash="dot", line_color="#6b7280", line_width=1)

if payback != float('inf') and payback <= meses_total:
    fig.add_vline(x=payback, line_dash="dot", line_color="#fb923c", line_width=1.5,
                  annotation_text=f"Payback: {payback_texto}",
                  annotation_font_color="#fb923c", annotation_position="top right")

fig.update_layout(
    paper_bgcolor="#0a0a0f", plot_bgcolor="#0f0f1a",
    font=dict(family="Space Mono, monospace", color="#9ca3af", size=11),
    legend=dict(bgcolor="rgba(26,26,46,0.9)", bordercolor="#2d2d4e", borderwidth=1, font=dict(size=11)),
    xaxis=dict(title="Meses", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e"),
    yaxis=dict(title="R$", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e", tickprefix="R$ ", tickformat=",.0f"),
    hovermode="x unified", margin=dict(l=10, r=10, t=20, b=10), height=380,
)
st.plotly_chart(fig, use_container_width=True)

# ── Tabela expansível ──────────────────────────────────────────────────────────
with st.expander("📋 Ver tabela de projeção mês a mês"):
    df_proj = pd.DataFrame({
        "Mês": meses_eixo,
        "Custo Acumulado": [f"R$ {v:,.2f}" for v in custo_acum],
        "Benefício Acumulado": [f"R$ {v:,.2f}" for v in benef_acum],
        "Saldo Líquido": [f"R$ {v:,.2f}" for v in saldo_acum],
        "ROI Acumulado": [f"{((b-c)/c*100):.1f}%" if c > 0 else "-" for b, c in zip(benef_acum, custo_acum)],
    })
    st.dataframe(df_proj, use_container_width=True, hide_index=True)

# ── Resumo da Viabilidade (Texto Atualizado) ───────────────────────────────────
st.markdown(f"""
<div class="summary-box">
<h3 style="color: #4ade80; font-size: 16px; margin-top: 0; margin-bottom: 12px;">Resumo da Viabilidade</h3>
<p>
A manutenção de <strong>R$ {custo_manut:,.2f}</strong> representa apenas <strong>{perc_manut:.1f}%</strong> do benefício gerado (<strong>R$ {benef_mensal:,.2f}</strong>).<br><br>
Isso significa que a automação se paga "sozinha" e ainda sobra uma margem de segurança enorme. 
O Payback de <strong>{payback_texto}</strong> indica que, antes do período de retorno terminar, você já recuperou todo o dinheiro investido no desenvolvimento e na manutenção acumulada.
<br><br>
{veredicto_emoji}
</p>
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
st.caption("Desenvolvido Por Rodrigo Aiosa, Todos os Diretos Reservados")

